import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import tempfile
import os

def classify_day(x):
    if x < 5:
        return 'Weekday'
    elif x == 5:
        return 'Saturday'
    elif x == 6:
        return 'Sunday'

def process_files(product_sales_file, wastage_sales_file, forecast_file=None):
    # Columns we care about
    cols_needed = ["Outlet", "Item", "Business Date", "Net Sales", "Item Qty"]

    # --- Sales ---
    sales_df = pd.read_excel(product_sales_file, header=6)

    # Force first two columns to be 'Outlet' and 'Item'
    sales_df.columns.values[0] = "Outlet"
    sales_df.columns.values[1] = "Item"

    # Keep only the needed columns (ignore extras)
    sales_df = sales_df[[c for c in cols_needed if c in sales_df.columns]]

    sales_df['Outlet'] = sales_df['Outlet'].ffill()
    sales_df['Item'] = sales_df['Item'].ffill()
    sales_df = sales_df[~sales_df['Business Date'].isin(['Subtotal', 'Grand Total', 'NaN'])]
    sales_df = sales_df.dropna()

    # --- Wastage ---
    wastage_df = pd.read_excel(wastage_sales_file, header=6)

    # Force first two columns to be 'Outlet' and 'Item'
    wastage_df.columns.values[0] = "Outlet"
    wastage_df.columns.values[1] = "Item"

    # Keep only the needed columns (ignore extras)
    wastage_df = wastage_df[[c for c in cols_needed if c in wastage_df.columns]]

    wastage_df['Outlet'] = wastage_df['Outlet'].ffill()
    wastage_df['Item'] = wastage_df['Item'].ffill()
    wastage_df = wastage_df[~wastage_df['Business Date'].isin(['Subtotal', 'Grand Total', 'NaN'])]
    wastage_df = wastage_df.dropna()


    # Merge
    merged_df = pd.merge(sales_df, wastage_df, on=["Outlet", "Item", "Business Date"], how="outer")
    merged_df.rename(columns={
        "Item Qty_x": "Qty",
        "Net Sales_x": "Sales",
        "Item Qty_y": "Wastage Qty",
        "Net Sales_y": "Wastage Sales"
    }, inplace=True)

    merged_df['Business Date'] = pd.to_datetime(merged_df['Business Date'], errors='coerce')
    merged_df['day_of_week'] = merged_df['Business Date'].dt.dayofweek
    merged_df['Day Type'] = merged_df['day_of_week'].apply(classify_day)

    # Group averages
    avg_data = (
        merged_df.groupby(['Outlet', 'Item', 'Day Type'])[
            ['Sales', 'Qty', 'Wastage Sales', 'Wastage Qty']
        ]
        .mean()
        .reset_index()
    )


    avg_data_long = avg_data.copy()
    avg_data_long['Item'] = avg_data_long['Item'].astype(str).str.split('-', n=1).str[1]
    avg_data_long['Outlet'] = avg_data_long['Outlet'].astype(str).str.split('-KOMUGI', n=1).str[1]

    # Reorder columns
    desired_order = ['Outlet', 'Item', 'Day Type', 'Qty', 'Wastage Qty', 'Sales', 'Wastage Sales']
    avg_data_long = avg_data_long[desired_order]

    # Round numeric columns
    numeric_cols_0 = ['Qty', 'Wastage Qty']
    for col in numeric_cols_0:
        nums = pd.to_numeric(avg_data_long[col], errors='coerce')
        rounded = nums.round(0)
        avg_data_long[col] = np.where(nums.notna(), rounded, avg_data_long[col])

    numeric_cols_2 = ['Sales', 'Wastage Sales']
    for col in numeric_cols_2:
        nums = pd.to_numeric(avg_data_long[col], errors='coerce')
        rounded = nums.round(2)
        avg_data_long[col] = np.where(nums.notna(), rounded, avg_data_long[col])

    # Reorder day type
    day_order = pd.CategoricalDtype(categories=["Weekday", "Saturday", "Sunday"], ordered=True)
    avg_data_long['Day Type'] = avg_data_long['Day Type'].astype(day_order)
    avg_data_long = avg_data_long.sort_values(by=['Outlet', 'Item', 'Day Type']).reset_index(drop=True)
    avg_data_long["Wastage Qty"] = avg_data_long["Wastage Qty"].fillna(0)

    # Compute total averages
    item_totals = (
        avg_data_long.groupby(['Outlet', 'Item'])['Qty']
        .mean()
        .reset_index()
        .rename(columns={'Qty': 'Total Avg Qty'})
    )
    item_totals['Total Avg Qty'] = pd.to_numeric(item_totals['Total Avg Qty'], errors='coerce')

    top_bottom = {}
    for outlet, group in item_totals.groupby('Outlet'):
        top_items = group.nlargest(10, 'Total Avg Qty')['Item']
        bottom_items = group.nsmallest(10, 'Total Avg Qty')['Item']
        top_bottom[outlet] = {'top': set(top_items), 'bottom': set(bottom_items)}

    # Save to Excel in a temporary file
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    output_file_path = tmp_file.name

    with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:

        if forecast_file is None:
            for outlet, group in avg_data_long.groupby("Outlet"):
                sheet_name = str(outlet).replace("Outlet: ", "")[:31]
                group = group.drop(columns=["Outlet"])
                group.to_excel(writer, sheet_name=sheet_name, index=False)

        # Add MV Forecast Comparison if forecast file is provided
        else:
            # forecast_mv = pd.read_excel(forecast_file, sheet_name="MV")
            outlet_map = {
                "MV": "Mid Valley",
                "PV": "Pavilion",
                "OU": "One Utama",
                "SA": "AEON Shah Alam",
                "QM": "Quayside Mall",
                "MM": "Melawati Mall",
                "KLE": "KL East Mall",
                "KL": "Kuchai",
                "DP": "Dpulze",
                "SS2": "SS2",
                "PD": "Paradigm",
                "TP": "Taipan",
                "MP": "Main Place",
                "SW": "Sunway Pyramid",
            }

            for code, name in outlet_map.items():
                forecast_df = pd.read_excel(forecast_file, sheet_name=code)
                if code in ['MV', 'SA', 'QM']:
                    forecast_df = forecast_df.rename(columns={
                        "Item Name": "Item",
                        "Mon - Thu": "Weekday",
                        "Fri": "Weekday2",
                        "Sat": "Saturday",
                        "Sun": "Sunday"
                    })
                    forecast_df["Weekday"] = forecast_df[["Weekday", "Weekday2"]].mean(axis=1)
                else:
                    forecast_df = forecast_df.rename(columns={
                        "Item Name": "Item",
                        "Mon - Fri": "Weekday",
                        "Sat": "Saturday",
                        "Sun": "Sunday"
                    })

                forecast_long = forecast_df.melt(
                    id_vars=["Item"],
                    value_vars=["Weekday", "Saturday", "Sunday"],
                    var_name="Day Type",
                    value_name="Forecast Qty"
                )

                actual_df = avg_data_long[avg_data_long["Outlet"].str.contains(name, case=False)].copy()
                actual_df["Total"] = actual_df["Qty"] + actual_df["Wastage Qty"]

                comparison = pd.merge(
                    actual_df,
                    forecast_long,
                    on=["Item", "Day Type"],
                    how="left"
                )
                comparison["Variance"] = comparison["Total"] - comparison["Forecast Qty"]
                # comparison["% Variance"] = (comparison["Variance"] / comparison["Forecast Qty"]) * 100
                comparison = comparison.drop(columns=["Outlet", "Sales", "Wastage Sales"])

                # Add Recommendation column based on conditions
                def recommend(row):
                    if (row["Variance"] <= -10) and (row["Wastage Qty"] <= 5):
                        return "Increase Production (low wastage, under forecast)"
                    elif (row["Variance"] >= 10) and (row["Wastage Qty"] >= 5):
                        return "Decrease Production (high wastage, over forecast)"
                    else:
                        return "OK"

                comparison["Recommendation"] = comparison.apply(recommend, axis=1)

                # Keep Variance and % Variance numeric and rounded (no string '+')
                comparison["Variance"] = pd.to_numeric(comparison["Variance"], errors="coerce").round(0)
                # comparison["% Variance"] = pd.to_numeric(comparison["% Variance"], errors="coerce").round(0)

                outlet_name = name
                comparison.to_excel(writer, sheet_name=outlet_name, index=False)

    # Reopen and format
    wb = load_workbook(output_file_path)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for ws in wb.worksheets:

        col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        # Apply currency formatting to Sales & Wastage Sales if present
        for col_name in ["Sales", "Wastage Sales"]:
            if col_name in col_map:
                col_idx = col_map[col_name]
                for row in range(2, ws.max_row + 1):  # skip header
                    cell = ws.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = u'"RM"#,##0.00'
        # Find Item column
        item_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Item":
                item_col = idx
                break
        if not item_col:
            continue

        max_row = ws.max_row
        merge_start = 2
        current_val = ws.cell(row=merge_start, column=item_col).value

        for row in range(3, max_row + 2):  # sentinel
            val = ws.cell(row=row, column=item_col).value if row <= max_row else None
            if val != current_val:
                if row - merge_start > 1:
                    ws.merge_cells(start_row=merge_start, start_column=item_col,
                                   end_row=row - 1, end_column=item_col)
                ws.cell(row=merge_start, column=item_col).alignment = Alignment(vertical="center", horizontal="center")
                merge_start = row
                current_val = val

        # Highlight top/bottom
        outlet_name = ws.title
        if outlet_name in top_bottom:
            for row in range(2, max_row + 1):
                item_val = ws.cell(row=row, column=item_col).value
                if item_val in top_bottom[outlet_name]['top']:
                    ws.cell(row=row, column=item_col).fill = green_fill
                elif item_val in top_bottom[outlet_name]['bottom']:
                    ws.cell(row=row, column=item_col).fill = red_fill

        # Auto-adjust width with extra space for currency columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            header = col[0].value if col[0].value else ""
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except:
                    pass
            # Add extra padding for Sales columns so RM formatting fits
            if header in ["Sales"]:
                ws.column_dimensions[col_letter].width = max_length + 8
            elif header in ["Qty"]:
                ws.column_dimensions[col_letter].width = max_length + 6
            else:
                ws.column_dimensions[col_letter].width = max_length + 2

    # Apply conditional fills and number formats for Variance / % Variance in 'MV Forecast Comparison'
    if forecast_file is not None:
        for sheet_name in wb.sheetnames:
            # if not sheet_name.endswith("Comparison"):
            #     continue
            # print(sheet_name)
            ws_fc = wb[sheet_name]
            header_row = ws_fc[1]

            # Locate columns
            variance_col = None
            percent_col = None
            for idx, cell in enumerate(header_row, 1):
                if cell.value == "Variance":
                    variance_col = idx
                # elif cell.value == "% Variance":
                #     percent_col = idx

            # Number formats with explicit '+' for positives
            if variance_col:
                for row in range(2, ws_fc.max_row + 1):
                    c = ws_fc.cell(row=row, column=variance_col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '+#,##0;-#,##0;0'
                        # Color code by absolute variance thresholds
                        if c.value >=10:
                            c.fill = green_fill
                        elif c.value <= -10:
                            c.fill = red_fill

            # if percent_col:
            #     for row in range(2, ws_fc.max_row + 1):
            #         c = ws_fc.cell(row=row, column=percent_col)
            #         if isinstance(c.value, (int, float)):
            #             c.number_format = '+0"%" ;-0"%" ;0"%"'

    wb.save(output_file_path)
    return output_file_path

# ---------------- Streamlit App ----------------
st.title("🍞 Daily Sales Report Generator")

st.write("Upload your **Product Sales** and **Wastage Sales** Excel files:")

product_file = st.file_uploader("Product Sales File", type="xlsx")
wastage_file = st.file_uploader("Wastage Sales File", type="xlsx")
forecast_file = st.file_uploader("Forecast File (optional)", type="xlsx")


if st.button("Generate Report"):
    if product_file and wastage_file:
        try:
            with st.spinner("Processing files..."):
                output_path = process_files(product_file, wastage_file, forecast_file)
            with open(output_path, "rb") as f:
                st.download_button("⬇️ Download Report", f, file_name="avg_sales_by_outlet.xlsx")
            st.success("Report generated successfully!")
        except Exception as e:
            st.error(f"Error: {e}")
    else:
        st.error("Please upload both files first.")